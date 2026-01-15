from __future__ import annotations

import importlib.util
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ====== ŚCIEŻKI WZGLĘDNE DO STRUKTURY PROJEKTU ======
WORKCELL_DIR = Path(__file__).parent          # Workcell/
PROJECT_ROOT = WORKCELL_DIR.parent            # root projektu
OUTPUT_XLSX = PROJECT_ROOT / "Koszty_template.xlsx"

sys.path.insert(0, str(PROJECT_ROOT))

from Workcell._workcell_definitions import (
    Workcell,
    ParamChoice,
    ParamNumber,
    ParamTablePick,
)

# ====== STYL ======
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD_FONT = Font(bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True)

PLACEHOLDER_ROWS = 20


# ---------- helpers ----------

def _safe_sheet_name(name: str) -> str:
    return re.sub(r"[:\\/?*\[\]]", "_", name)[:31]


def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(max_len + 2, 50))


def _import_module(path: Path) -> Any:
    spec = importlib.util.spec_from_file_location(path.stem, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)  # type: ignore
    return mod


def _find_workcells(mod) -> list[Workcell]:
    return [v for v in vars(mod).values() if isinstance(v, Workcell)]


# ---------- główna funkcja ----------

def generate_cost_template() -> Path:
    """
    Generuje Koszty_template.xlsx na podstawie definicji gniazd.
    Zwraca ścieżkę do wygenerowanego pliku.
    """
    wb = Workbook()
    index_rows = []

    for py in WORKCELL_DIR.glob("*.py"):
        if py.name.startswith("_"):
            continue

        mod = _import_module(py)
        for wc in _find_workcells(mod):
            _generate_workcell_sheet(wb, wc)
            index_rows.append(_index_row(wc, py.name))

    _generate_index_sheet(wb, index_rows)
    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX


# ---------- generowanie arkuszy ----------

def _generate_workcell_sheet(wb: Workbook, wc: Workcell):

    choice = wc.choice_param
    if choice:
        choice_key = choice.key
        choice_values = choice.options
    else:
        choice_key = "DEFAULT"
        choice_values = ["DEFAULT"]

    # ========= NOWY WARUNEK =========
    def has_non_table_params() -> bool:
        for cv in choice_values:
            for formula in wc.formulas:
                if not _is_formula_allowed_for_choice(formula, choice_key, cv):
                    continue
                for p in formula.params:
                    if isinstance(p, ParamNumber) and p.source == "cost_table":
                        return True
        return False

    create_main_ws = has_non_table_params()
    ws = None
    row = 1

    if create_main_ws:
        ws = wb.create_sheet(_safe_sheet_name(f"WC__{wc.name}"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=f"{wc.name} (id={wc.id})").font = Font(bold=True, size=14)
        row += 2
    # ========= KONIEC WARUNKU =========

    # Tworzymy tabele dla każdego choice i każdej formuły
    for cv in choice_values:

        if ws is not None:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row=row, column=1, value=f"{choice_key} = {cv}").font = SECTION_FONT
            row += 1

        for formula in wc.formulas:
            if not _is_formula_allowed_for_choice(formula, choice_key, cv):
                continue

            # ---------- ParamNumber -> tylko jeśli arkusz główny istnieje ----------
            number_params = [
                p for p in formula.params
                if isinstance(p, ParamNumber) and p.source == "cost_table"
            ]

            if ws is not None and number_params:
                ws.append([f"Formula: {formula.id} - {formula.label}", "", "", ""])
                ws.cell(row=row, column=1).font = BOLD_FONT
                row += 1

                ws.append(["param_key", "label", "unit", "value"])
                for c in range(1, 5):
                    ws.cell(row=row, column=c).fill = HEADER_FILL
                    ws.cell(row=row, column=c).font = HEADER_FONT
                row += 1

                for p in number_params:
                    ws.append([p.key, p.label, getattr(p, "unit", ""), ""])
                    row += 1

                row += 1

            # ---------- ParamTablePick -> osobne arkusze ----------
            for tp in formula.params:
                if isinstance(tp, ParamTablePick) and tp.source == "cost_table":
                    tname = _tablepick_sheet_name(wc, tp)
                    tws = _get_or_create_sheet(wb, tname)

                    if tws.max_row == 1 and tws.cell(1, 1).value is None:
                        tws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
                        tws.cell(
                            row=1,
                            column=1,
                            value=f"{wc.name} / {tp.table_name}"
                        ).font = Font(bold=True, size=14)
                        tws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
                        next_row = 3
                    else:
                        next_row = tws.max_row + 2

                    block_title = f"{choice_key}={cv} | formula={formula.id} ({formula.label})"
                    next_row = _append_tablepick_block(
                        tws, block_title, tp, next_row, PLACEHOLDER_ROWS
                    )

                    _autosize(tws)

        if ws is not None:
            row += 1

    if ws is not None:
        _autosize(ws)



def _is_formula_allowed_for_choice(formula, choice_key: str, choice_value: str) -> bool:
    """
    Zwraca True jeśli formuła jest dozwolona dla (choice_key=choice_value).
    Jeśli brak choice_key (DEFAULT) – traktujemy formułę jako dozwoloną, o ile nie ma reguł wyboru.
    Jeśli enabled_when puste -> dozwolone zawsze.
    """

    rules = list(getattr(formula, "enabled_when", []) or [])

    # 1) brak warunków -> dozwolone zawsze
    if not rules:
        return True

    # 2) jeżeli ten workcell nie ma choice_param -> choice_key to "DEFAULT"
    #    a reguły odnoszące się do choice_key nie mogą być spełnione, więc:
    #    - jeśli formuła ma jakiekolwiek reguły na choice_key, to nie jest dozwolona w DEFAULT
    if choice_key == "DEFAULT":
        for r in rules:
            if getattr(r, "param_key", None) == choice_key:
                return False
        # brak reguł na choice -> traktuj jako dozwoloną
        return True

    # 3) normalny przypadek: mamy choice_key (np typ_ocynku)
    #    Jeśli istnieją reguły na choice_key, to muszą pasować do choice_value
    has_choice_rule = False
    for r in rules:
        if getattr(r, "param_key", None) == choice_key:
            has_choice_rule = True
            if getattr(r, "equals", None) != choice_value:
                return False

    # jeśli były reguły na choice_key i żadna nie odrzuciła -> OK
    # jeśli nie było reguł na choice_key -> uznajemy, że formuła dostępna dla wszystkich choice
    return True

def _index_row(wc: Workcell, filename: str):
    return [
        wc.id,
        wc.name,
        filename,
    ]


def _generate_index_sheet(wb: Workbook, rows):
    ws = wb.active
    ws.title = "INDEX"
    ws.append(["workcell_id", "name", "source_file"])
    for c in range(1, 4):
        ws.cell(row=1, column=c).fill = HEADER_FILL
        ws.cell(row=1, column=c).font = HEADER_FONT

    for r in rows:
        ws.append(r)

    _autosize(ws)

def _tablepick_sheet_name(wc: Workcell, tp: ParamTablePick) -> str:
    # wymaganie: WC_<gniazdo>_<param_key>
    return _safe_sheet_name(f"WC_{wc.name}_{tp.key}")

def _get_or_create_sheet(wb: Workbook, title: str):
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title)

def _append_name_value_block(ws, title: str, start_row: int, placeholder_rows: int = PLACEHOLDER_ROWS) -> int:
    """
    Dodaje blok:
      <title> (merged)
      Nazwa | Wartość (header)
      <placeholder_rows> pustych wierszy
    Zwraca nowy row pointer.
    """
    row = start_row

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    ws.append(["Nazwa", "Wartość"])
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).font = HEADER_FONT
    row += 1

    # placeholdery
    for _ in range(placeholder_rows):
        ws.append(["", ""])
        row += 1

    # odstęp
    row += 1
    return row


def _workcell_has_non_table_cost_params(wc: Workcell, choice_key: str, choice_values: list[str]) -> bool:
    """
    True jeśli istnieje przynajmniej jeden ParamNumber(source="cost_table")
    w jakiejkolwiek formule dozwolonej dla jakiegokolwiek choice.
    """
    for cv in choice_values:
        for formula in wc.formulas:
            if not _is_formula_allowed_for_choice(formula, choice_key, cv):
                continue
            for p in formula.params:
                if isinstance(p, ParamNumber) and p.source == "cost_table":
                    return True
    return False

def _append_tablepick_block(
    ws,
    title: str,
    tp: ParamTablePick,
    start_row: int,
    placeholder_rows: int = PLACEHOLDER_ROWS,
) -> int:
    """
    Dodaje blok tabeli ParamTablePick:
      <title> (merged)
      <table_name> | <label> <unit>
      <placeholder_rows> pustych wierszy
    Zwraca nowy row pointer.
    """
    row = start_row

    # tytuł bloku
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    # nagłówek tabeli
    col1 = tp.table_name
    col2 = f"{tp.label} {tp.unit}".strip()

    ws.append([col1, col2])
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).font = HEADER_FONT
    row += 1

    # placeholdery
    for _ in range(placeholder_rows):
        ws.append(["", ""])
        row += 1

    # odstęp
    row += 1
    return row
