from __future__ import annotations

import importlib.util
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from Workcell._workcell_definitions import Workcell, ParamNumber, ParamTablePick


WORKCELL_DIR = Path(__file__).parent
PROJECT_ROOT = WORKCELL_DIR.parent


@dataclass
class CostData:
    # wc_id -> choice -> formula_id -> param_key -> value
    numbers: dict[int, dict[str, dict[str, dict[str, float]]]]
    # wc_id -> choice -> formula_id -> table_param_key -> item_name -> value
    tables: dict[int, dict[str, dict[str, dict[str, dict[str, float]]]]]

    def get_number(self, wc_id: int, choice: str, formula_id: str, key: str) -> float:
        return self.numbers[wc_id][choice][formula_id][key]

    def get_table_value(self, wc_id: int, choice: str, formula_id: str, table_param_key: str, item_name: str) -> float:
        return self.tables[wc_id][choice][formula_id][table_param_key][item_name]


def _import_module(path: Path) -> Any:
    spec = importlib.util.spec_from_file_location(path.stem, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot import: {path}")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)  # type: ignore[attr-defined]
    return mod


def _find_workcells(mod: Any) -> list[Workcell]:
    return [v for v in vars(mod).values() if isinstance(v, Workcell)]


def load_workcell_definitions() -> dict[int, Workcell]:
    """
    Skanuje Workcell/*.py (bez tych od '_') i zwraca mapę: workcell_id -> Workcell
    """
    out: dict[int, Workcell] = {}
    for py in WORKCELL_DIR.glob("*.py"):
        if py.name.startswith("_"):
            continue
        mod = _import_module(py)
        for wc in _find_workcells(mod):
            if wc.id in out:
                raise ValueError(f"Duplicate workcell id={wc.id} ({wc.name}) in {py.name}")
            out[wc.id] = wc
    return out


def _to_float(v: Any) -> float:
    if v is None:
        raise ValueError("Empty cost value")
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            raise ValueError("Empty cost value")
        # PL: przecinek dziesiętny + spacje tysięcy
        s = s.replace(" ", "").replace("\u00A0", "").replace(",", ".")
        return float(s)
    raise ValueError(f"Unsupported value type: {type(v)} ({v!r})")


def load_costs_from_excel(xlsx_path: Path, workcells: dict[int, Workcell]) -> CostData:
    """
    Czyta Koszty_template.xlsx wygenerowany nowym generatorem:
    - ParamNumber(cost_table) są w arkuszach WC__<Gniazdo> (o ile istnieją)
    - ParamTablePick(cost_table) są w arkuszach WC_<Gniazdo>_<param_key>
      i mają bloki: "choice=... | formula=..." + tabela 2 kolumny (item_name | value)
    """
    wb = load_workbook(xlsx_path, data_only=True)

    numbers: dict[int, dict[str, dict[str, dict[str, float]]]] = {}
    tables: dict[int, dict[str, dict[str, dict[str, dict[str, float]]]]] = {}

    if "INDEX" not in wb.sheetnames:
        raise ValueError("Excel missing INDEX sheet")

    # INDEX: workcell_id, name, source_file
    index_ws = wb["INDEX"]
    name_to_id: dict[str, int] = {}
    for row in index_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        name_to_id[str(row[1])] = int(row[0])

    # pomocnicze: lista nazw gniazd do dopasowania w arkuszach WC_<gniazdo>_<param_key>
    wc_names_sorted = sorted(name_to_id.keys(), key=len, reverse=True)

    def ensure(wc_id: int, choice: str, formula_id: str):
        numbers.setdefault(wc_id, {}).setdefault(choice, {}).setdefault(formula_id, {})
        tables.setdefault(wc_id, {}).setdefault(choice, {}).setdefault(formula_id, {})

    def is_blank(v: Any) -> bool:
        return v is None or (isinstance(v, str) and not v.strip())

    # -------------------------
    # A) WC__<Gniazdo> (ParamNumber)
    # -------------------------
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("WC__"):
            continue

        ws = wb[sheet_name]
        wc_name = sheet_name.replace("WC__", "", 1)

        if wc_name not in name_to_id:
            raise ValueError(f"Sheet {sheet_name} not found in INDEX (workcell name mismatch)")
        wc_id = name_to_id[wc_name]

        wc_def = workcells.get(wc_id)
        if wc_def is None:
            raise ValueError(f"Workcell id={wc_id} from INDEX not found in definitions")

        if wc_def.choice_param:
            choice_key = wc_def.choice_param.key
            choice_values = list(wc_def.choice_param.options or [])
        else:
            choice_key = "DEFAULT"
            choice_values = ["DEFAULT"]

        current_choice: str | None = None
        current_formula_id: str | None = None

        i = 1
        max_row = ws.max_row

        while i <= max_row:
            a = ws.cell(row=i, column=1).value

            # choice line: "<choice_key> = <value>"
            if isinstance(a, str) and a.strip().startswith(f"{choice_key} ="):
                current_choice = a.split("=", 1)[1].strip()
                if current_choice not in choice_values:
                    raise ValueError(f"{sheet_name}: unknown choice '{current_choice}' (expected {choice_values})")
                current_formula_id = None
                i += 1
                continue

            # formula line: "Formula: <id> - <label>"
            if isinstance(a, str) and a.strip().startswith("Formula:"):
                s = a.strip()[len("Formula:"):].strip()
                current_formula_id = s.split(" - ", 1)[0].strip() if " - " in s else s.strip()
                if current_choice is None:
                    current_choice = "DEFAULT"
                ensure(wc_id, current_choice, current_formula_id)
                i += 1
                continue

            # header row: param_key ... value (col4)
            if (
                current_choice is not None
                and current_formula_id is not None
                and isinstance(a, str)
                and a.strip() == "param_key"
                and str(ws.cell(row=i, column=4).value).strip() == "value"
            ):
                i += 1
                while i <= max_row:
                    pk = ws.cell(row=i, column=1).value
                    if is_blank(pk):
                        break
                    pk_s = str(pk).strip()

                    val = ws.cell(row=i, column=4).value
                    if is_blank(val):
                        raise ValueError(f"{sheet_name} ({current_choice}) formula={current_formula_id}: missing value for {pk_s}")

                    numbers[wc_id][current_choice][current_formula_id][pk_s] = _to_float(val)
                    i += 1
                continue

            i += 1

    # -------------------------
    # B) WC_<Gniazdo>_<param_key> (ParamTablePick)
    # -------------------------
    for sheet_name in wb.sheetnames:
        if sheet_name == "INDEX" or sheet_name.startswith("WC__"):
            continue
        if not sheet_name.startswith("WC_"):
            continue

        ws = wb[sheet_name]
        rest = sheet_name[len("WC_"):]

        wc_name = None
        table_param_key = None
        for cand in wc_names_sorted:
            prefix = f"{cand}_"
            if rest.startswith(prefix):
                wc_name = cand
                table_param_key = rest[len(prefix):]
                break
        if wc_name is None or table_param_key is None:
            raise ValueError(f"Cannot parse TablePick sheet name: {sheet_name}")

        wc_id = name_to_id[wc_name]
        wc_def = workcells.get(wc_id)
        if wc_def is None:
            raise ValueError(f"Workcell id={wc_id} not found in definitions")

        if wc_def.choice_param:
            choice_key = wc_def.choice_param.key
            choice_values = list(wc_def.choice_param.options or [])
        else:
            choice_key = "DEFAULT"
            choice_values = ["DEFAULT"]

        i = 1
        max_row = ws.max_row

        while i <= max_row:
            a = ws.cell(row=i, column=1).value

            # tytuł bloku: "<choice_key>=<cv> | formula=<id> (..)"
            if isinstance(a, str) and "|" in a and "formula=" in a:
                title = a.strip()

                left = title.split("|", 1)[0].strip()  # "typ=OGNIOWY"
                if "=" not in left:
                    raise ValueError(f"{sheet_name}: bad block title (no '='): {title}")
                ck, cv = [x.strip() for x in left.split("=", 1)]
                if ck != choice_key:
                    raise ValueError(f"{sheet_name}: bad choice key '{ck}', expected '{choice_key}'")
                if cv not in choice_values:
                    raise ValueError(f"{sheet_name}: unknown choice '{cv}' (expected {choice_values})")

                right = title.split("|", 1)[1].strip()  # "formula=by_dm2 (...)"
                if not right.startswith("formula="):
                    raise ValueError(f"{sheet_name}: bad block title (no 'formula='): {title}")
                fid_part = right[len("formula="):].strip()
                formula_id = fid_part.split(" (", 1)[0].strip() if " (" in fid_part else fid_part.split()[0].strip()

                ensure(wc_id, cv, formula_id)

                # następny wiersz: nagłówek 2 kolumny (nie musimy walidować treści)
                i += 2

                tables[wc_id][cv][formula_id].setdefault(table_param_key, {})

                while i <= max_row:
                    item_name = ws.cell(row=i, column=1).value
                    val = ws.cell(row=i, column=2).value
                    if is_blank(item_name):
                        break
                    if is_blank(val):
                        raise ValueError(f"{sheet_name} ({cv}) formula={formula_id} table={table_param_key}: missing value for '{item_name}'")

                    tables[wc_id][cv][formula_id][table_param_key][str(item_name).strip()] = _to_float(val)
                    i += 1

                continue

            i += 1

    return CostData(numbers=numbers, tables=tables)


def load_workcells_with_costs():
    workcells = load_workcell_definitions()
    xlsx_path = PROJECT_ROOT / "Koszty_template.xlsx"

    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"Brak pliku kosztów: {xlsx_path}\n"
            f"Najpierw wygeneruj template (_generate_cost_template.py)."
        )

    costs = load_costs_from_excel(xlsx_path, workcells)
    return workcells, costs
